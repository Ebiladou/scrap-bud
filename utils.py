import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("School", 30),
    ("Society Name", 35),
    ("President", 25),
    ("Email", 35),
    ("Society URL", 55),
]

def _make_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def save_to_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Societies"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    cell_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="DCE6F1")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = _make_border()

    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [row["school"], row["name"], row["president"], row["email"], row["url"]]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = border
            if col_idx == 5 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            else:
                cell.font = cell_font
            cell.alignment = left_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

    wb.save(output_path)